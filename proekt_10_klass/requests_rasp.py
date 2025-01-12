import urllib.request
import requests
import urllib
import openpyxl

class ClassOfGetRasp:
    def id_file_of_rasp(self):
        response = requests.get("""https://sites.google.com/view/perspectiva99/%D0%B3%D0%BB%D0%B0%D0%B2%D0%BD%D0%B0%D1%8F-%D1%81%D1%82%D1%80%D0%B0%D0%BD%D0%B8%D1%86%D0%B0""")
        response = response.text
        response = response
        index = response.find("""data-embed-doc-id""")

        response= response[index:index+100]
        response = response[response.find('"') + 1:index+100]
        result = response[:response.find('"')]
        return result

    def download_rasp(self):
        try:
            file_xlsx = urllib.request.urlopen(f"https://docs.google.com/spreadsheets/d/{self.id_file_of_rasp()}/export?format=xlsx").read()
            # file_xlsx = urllib.request.urlopen(f"https://drive.usercontent.google.com/u/0/uc?id={id_file_of_rasp()}&export=download").read()
            with open(file='расписание.xlsx', mode='wb') as file:
                file.write(file_xlsx)
                file.close()
            return True
        except:
            return False

    def return_rasp_for_user(self, grade):
        try:
            grade = float(grade)
        except:
            None

        xlsx = openpyxl.open('расписание.xlsx', data_only=True).active

        all_cols = list(xlsx.iter_cols(values_only=True))
        all_rows = list(xlsx.iter_rows(values_only=True))

        # if grade in all_cols[0]:
        #     return False

        col = ()

        for i in all_cols:
            if grade in i:
                col = i

        max_lessons = 0

        try:
            for i in all_cols[0][col.index(grade)+1:]:
                try:
                    max_lessons = int(i)
                except:
                    break
        except:
            return False

        col = list(col[col.index(grade)+1:col.index(grade)+max_lessons+1])

        for i in col[::-1]:
            if i == None:
                col.pop(len(col)-1)
            else:
                break

        for i in range(0, len(col)-1):
            if col[i] == None:
                col[i] = 'Окно'

        for i in range(0, len(col)):
            col[i] = f'{i+1}) {col[i]}'

        res_lessons = ''

        for i in col:
            res_lessons += f'{i}\n'

        try:
            grade = str(int(grade))
        except:
            None
            
        try:
            date_time = str(all_rows[0][0].date())
        except:
            date_time = str(all_rows[0][0])
        
        return date_time + '\n\n' + grade + '\n' + res_lessons

    def write_all_classes(self):
        xlsx = openpyxl.open('расписание.xlsx').active

        all_classes = []

        all_rows = list(xlsx.iter_rows(values_only=True))

        for row in all_rows:
            if row[0] == '#':
                for i in row[1:]:
                    if i != None:
                        all_classes.append(i)

        with open(file = 'all_classes.txt', mode='w+', encoding='utf-8') as file:
            for i in all_classes:
                try:
                    i = int(i)
                except:
                    None

                file.write(f'{i}\n')


    def return_all_classes(self):

        with open(file = 'all_classes.txt', mode='r', encoding='utf-8') as file:
            all_classes = file.readlines()

        for i in range(0, len(all_classes)):
            all_classes[i] = all_classes[i][:all_classes[i].index('\n')]

        for i in range(0, len(all_classes)):
            try:
                all_classes[i] = int(all_classes[i])
            except:
                None
        
        return all_classes